"""Data router — economics, daily summary, fuel prices, sales, blackout, generators, site detail, upload."""
import os
import tempfile
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, Query, HTTPException
import pandas as pd
from utils.database import get_db
from models.energy_cost import get_store_economics, get_generator_detail, get_trends
from backend.routers.auth import get_current_user

router = APIRouter()


# ─── Helper: date filter SQL ─────────────────────────────
def _dsql(col: str, date_from: str = None, date_to: str = None):
    sql = ""
    params = []
    if date_from:
        sql += f" AND {col} >= ?"
        params.append(date_from)
    if date_to:
        sql += f" AND {col} <= ?"
        params.append(date_to)
    return sql, params


# ─── 1.6 Upload ──────────────────────────────────────────
@router.post("/upload/validate")
async def validate_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Validate file type without importing — returns detected type and sheets."""
    if user["role"] not in ("super_admin", "admin"):
        raise HTTPException(403, "Upload requires admin role")
    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        # Detect type by sheet names
        detected_type = None
        sheets_lower = [s.lower().strip() for s in sheets]

        if any(s in sheets_lower for s in ["daily sales", "daily_sales", "hourly sales", "hourly_sales"]):
            # Check row count to classify as daily vs reference
            try:
                import pandas as _pd
                for sn in sheets:
                    if "daily" in sn.lower() and "sales" in sn.lower():
                        row_count = len(_pd.read_excel(tmp_path, sheet_name=sn, nrows=15000))
                        detected_type = "sales_reference" if row_count > 10000 else "combo_sales"
                        break
                if not detected_type:
                    detected_type = "combo_sales"
            except Exception:
                detected_type = "combo_sales"
        elif len(sheets) >= 3 and sum(s in sheets_lower for s in ["cmhl", "cp", "cfc", "pg"]) >= 3:
            detected_type = "fuel_price"
        elif "cp" in sheets_lower:
            detected_type = "blackout_cp"
        elif "cmhl" in sheets_lower:
            detected_type = "blackout_cmhl"
        elif "cfc" in sheets_lower:
            detected_type = "blackout_cfc"
        elif "pg" in sheets_lower:
            detected_type = "blackout_pg"
        else:
            # Check content of first sheet for specific patterns
            try:
                import pandas as _pd
                _df = _pd.read_excel(tmp_path, sheet_name=0, nrows=5)
                all_text = ' '.join(str(c).lower() for c in _df.columns) + ' ' + ' '.join(str(v).lower() for v in _df.iloc[0].values if v is not None)

                if "daily average diesel" in all_text or "daily normal diesel" in all_text or "diesel expense" in all_text:
                    detected_type = "diesel_expense_ly"
                elif "daily sales" in all_text or "sales_amt" in all_text:
                    detected_type = "daily_sales"
                elif "store exp" in all_text or "expense percentage" in all_text or "store contribution" in all_text:
                    detected_type = "diesel_expense_ly"  # treat as reference data
                elif "gold_code" in all_text or "goldcode" in all_text or "store master" in all_text:
                    detected_type = "store_master"
            except Exception:
                pass

        if not detected_type:
            raise HTTPException(400, f"UNKNOWN_FILE_TYPE — sheets: {', '.join(sheets)}")

        # Estimate row count
        row_estimate = 0
        try:
            import pandas as _pd2
            for sn in sheets:
                sn_lower = sn.lower().strip()
                if sn_lower in ("cmhl", "cp", "cfc", "pg") or "daily" in sn_lower or "sales" in sn_lower or "hourly" in sn_lower:
                    df_peek = _pd2.read_excel(tmp_path, sheet_name=sn)
                    row_estimate += len(df_peek)
                elif sn_lower in ("sheet1", "data"):
                    df_peek = _pd2.read_excel(tmp_path, sheet_name=sn)
                    row_estimate += len(df_peek)
            if row_estimate == 0:
                df_peek = _pd2.read_excel(tmp_path, sheet_name=0)
                row_estimate = len(df_peek)
        except Exception:
            pass

        return {"type": detected_type, "sheets": sheets, "filename": file.filename, "rows": row_estimate}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"VALIDATION_ERROR: {e}")
    finally:
        os.unlink(tmp_path)


@router.post("/upload")
async def upload_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    if user["role"] not in ("super_admin", "admin"):
        raise HTTPException(403, "Upload requires admin role")

    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        sheets_lower = [s.lower().strip() for s in sheets]
        detected_type = None
        record_count = 0

        from utils.database import (
            get_db as _get_db, upsert_site, upsert_generator, upsert_daily_operation,
            insert_fuel_purchase, refresh_site_summary, log_upload,
            upsert_daily_sale, upsert_hourly_sale, upsert_diesel_expense_ly,
            enrich_sites_from_store_master,
        )

        # ─── FUEL PRICE (must check BEFORE blackout — both have sector sheets) ──
        if len(sheets) >= 3 and sum(s in sheets_lower for s in ["cmhl", "cp", "cfc", "pg"]) >= 3 and not any(s in sheets_lower for s in ["daily sales", "daily_sales"]):
            from parsers.fuel_price_parser import parse_fuel_price_file
            parsed = parse_fuel_price_file(tmp_path)
            purchases = parsed.get("purchases", []) if isinstance(parsed, dict) else []
            detected_type = "fuel_price"

            with _get_db() as conn:
                log_upload(conn, file.filename, "fuel_price", None,
                           len(purchases), len(purchases), 0, None, None, None)
                for fp in purchases:
                    insert_fuel_purchase(conn, fp["sector_id"], fp["date"],
                                         fp.get("region"), fp.get("supplier"),
                                         fp.get("fuel_type"), fp.get("quantity_liters"),
                                         fp.get("price_per_liter"), source="api")
            record_count = len(purchases)

        # ─── BLACKOUT FILES ────────────────────────────────────
        elif any(s in sheets_lower for s in ["cp", "cmhl", "cfc", "pg"]) and not any(s in sheets_lower for s in ["daily sales", "daily_sales"]):
            from parsers.blackout_parser import parse_blackout_file
            sector = next((s for s in ["CP", "CMHL", "CFC", "PG"] if s.lower() in sheets_lower), None)
            if sector:
                parsed = parse_blackout_file(tmp_path, sector)
                gens = parsed.get("generators", [])
                daily = parsed.get("daily_data", [])
                dates = parsed.get("dates_found", [])
                errors = parsed.get("errors", [])
                detected_type = f"blackout_{sector.lower()}"

                with _get_db() as conn:
                    batch_id = log_upload(conn, file.filename, detected_type, sector,
                                          len(daily), len(daily) - len(errors), len(errors),
                                          min(dates) if dates else None, max(dates) if dates else None,
                                          errors[:50] if errors else None)
                    gen_id_map = {}
                    for gen in gens:
                        upsert_site(conn, gen["site_id"], gen["site_name"], sector, gen["site_type"],
                                    cost_center_code=gen.get("cost_center_code"),
                                    business_sector=gen.get("business_sector"),
                                    company=gen.get("company"),
                                    site_code=gen.get("site_code"))
                        gen_id = upsert_generator(conn, gen["site_id"], gen["model_name"],
                                                   gen["model_name_raw"], gen["power_kva"],
                                                   gen["consumption_per_hour"], gen["fuel_type"], gen["supplier"])
                        gen_id_map[(gen["site_id"], gen["model_name_raw"])] = gen_id

                    sites_dates = set()
                    for row in daily:
                        gen_key = (row["site_id"], row["model_name_raw"])
                        gen_id = gen_id_map.get(gen_key)
                        if gen_id:
                            upsert_daily_operation(conn, gen_id, row["site_id"], row["date"],
                                                    row["gen_run_hr"], row["daily_used_liters"],
                                                    row["spare_tank_balance"], row["blackout_hr"],
                                                    source="api", upload_batch_id=batch_id)
                            sites_dates.add((row["site_id"], row["date"]))
                    for sid, dt in sites_dates:
                        refresh_site_summary(conn, sid, dt)

                    # Map sales site_id: since site_id = cost_center_code,
                    # sales_site_name (CostCenter) directly matches site_id
                    conn.execute("""
                        UPDATE daily_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = daily_sales.sales_site_name
                        )
                    """)
                    conn.execute("""
                        UPDATE hourly_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = hourly_sales.sales_site_name
                        )
                    """)

                    # Enrich sites with store master data (address, lat/long, etc.)
                    try:
                        enrich_sites_from_store_master(conn)
                    except Exception:
                        pass

                # ─── Build validation comparison: Excel vs Pipeline vs DB ───
                try:
                    import openpyxl as _xl
                    from parsers.base_parser import parse_date_from_cell as _pdc, clean_numeric as _cn
                    from collections import defaultdict as _dd

                    # 1. Excel totals from total row
                    _wb = _xl.load_workbook(tmp_path, data_only=True)
                    _ws = _wb[sector] if sector in _wb.sheetnames else _wb[_wb.sheetnames[0]]
                    _total_row = None
                    for _r in range(1, 15):
                        for _c in range(1, 8):
                            _v = _ws.cell(row=_r, column=_c).value
                            if _v and "total" in str(_v).lower():
                                _total_row = _r; break
                        if _total_row: break

                    _cd = parsed.get("columns_detected", {})
                    _dsub = _cd.get("date_sub_offsets", {})
                    _og, _of, _ot, _ob = _dsub.get("gen_run_hr",0), _dsub.get("daily_used_liters",1), _dsub.get("spare_tank_balance",2), _dsub.get("blackout_hr",3)
                    _dcols = []
                    for _ci in range(1, _ws.max_column + 1):
                        _ds = _pdc(_ws.cell(row=2, column=_ci).value)
                        if _ds: _dcols.append((_ds, _ci))

                    _excel = {}
                    if _total_row:
                        for _ds, _dc in _dcols:
                            _excel[_ds] = {
                                "gen_hr": _cn(_ws.cell(row=_total_row, column=_dc+_og).value) or 0,
                                "fuel": _cn(_ws.cell(row=_total_row, column=_dc+_of).value) or 0,
                                "tank": _cn(_ws.cell(row=_total_row, column=_dc+_ot).value) or 0,
                                "blackout": _cn(_ws.cell(row=_total_row, column=_dc+_ob).value) or 0,
                            }
                    _wb.close()

                    # 2. Parser totals
                    _ptot = _dd(lambda: {"gen_hr":0,"fuel":0,"tank":0,"blackout":0,"sites":set()})
                    for _row in daily:
                        _d = _row["date"]
                        _ptot[_d]["gen_hr"] += _row["gen_run_hr"] or 0
                        _ptot[_d]["fuel"] += _row["daily_used_liters"] or 0
                        _ptot[_d]["tank"] += _row["spare_tank_balance"] or 0
                        _ptot[_d]["blackout"] += _row["blackout_hr"] or 0
                        _ptot[_d]["sites"].add(_row["site_id"])

                    # 3. DB totals
                    import pandas as _pd
                    _dbdf = _pd.read_sql_query(f"""
                        SELECT dss.date, SUM(dss.total_gen_run_hr) as gen_hr, SUM(dss.total_daily_used) as fuel,
                               SUM(dss.spare_tank_balance) as tank, SUM(dss.blackout_hr) as blackout,
                               COUNT(DISTINCT dss.site_id) as sites
                        FROM daily_site_summary dss JOIN sites s ON s.site_id=dss.site_id
                        WHERE s.sector_id='{sector}' GROUP BY dss.date ORDER BY dss.date
                    """, conn)
                    _dbtot = {}
                    for _, _r in _dbdf.iterrows():
                        _dbtot[_r["date"]] = {"gen_hr":_r["gen_hr"] or 0,"fuel":_r["fuel"] or 0,"tank":_r["tank"] or 0,"blackout":_r["blackout"] or 0,"sites":int(_r["sites"])}

                    # 4. Build comparison rows
                    validation = []
                    _all_dates = sorted(set(list(_excel.keys()) + list(_ptot.keys()) + list(_dbtot.keys())))
                    for _d in _all_dates:
                        _e = _excel.get(_d, {"gen_hr":0,"fuel":0,"tank":0,"blackout":0})
                        _p = _ptot.get(_d, {"gen_hr":0,"fuel":0,"tank":0,"blackout":0,"sites":set()})
                        _db = _dbtot.get(_d, {"gen_hr":0,"fuel":0,"tank":0,"blackout":0,"sites":0})
                        _pass = all(abs(_e[k] - _db[k]) < 2 for k in ["gen_hr","fuel","tank","blackout"])
                        validation.append({
                            "date": _d,
                            "sites": len(_p["sites"]) if isinstance(_p["sites"], set) else 0,
                            "excel": {k: round(_e[k], 1) for k in ["gen_hr","fuel","tank","blackout"]},
                            "pipeline": {k: round(_p[k], 1) for k in ["gen_hr","fuel","tank","blackout"]},
                            "db": {k: round(_db[k], 1) for k in ["gen_hr","fuel","tank","blackout"]},
                            "pass": _pass,
                        })
                except Exception:
                    validation = []

                record_count = len(daily)

        # ─── COMBO SALES ───────────────────────────────────────
        elif any(s in sheets_lower for s in ["daily sales", "daily_sales", "hourly sales", "hourly_sales"]):
            from parsers.sales_parser import parse_daily_sales_file, parse_hourly_sales_file
            total = 0
            is_large = False

            # Daily sales
            for sn in sheets:
                if "daily" in sn.lower() and "sales" in sn.lower():
                    try:
                        result = parse_daily_sales_file(tmp_path, sheet_name=sn)
                        records = result.get("records", [])
                        is_large = len(records) > 10000

                        with _get_db() as conn:
                            ftype = "sales_reference" if is_large else "daily_sales"
                            batch_id = log_upload(conn, file.filename, ftype, None,
                                                  len(records), len(records), 0,
                                                  result.get("date_range", [None, None])[0],
                                                  result.get("date_range", [None, None])[1], None)

                            if is_large:
                                # Batch insert for large files (>10K rows)
                                conn.execute("BEGIN")
                                for i, r in enumerate(records):
                                    conn.execute("""
                                        INSERT OR REPLACE INTO daily_sales
                                        (sales_site_name, sector_id, date, brand, sales_amt, margin, source, upload_batch_id)
                                        VALUES (?, ?, ?, ?, ?, ?, 'api', ?)
                                    """, (r["sales_site_name"], r.get("sector_id"), r["date"],
                                          r["brand"], r["sales_amt"], r["margin"], batch_id))
                                conn.execute("COMMIT")
                            else:
                                for r in records:
                                    upsert_daily_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                                      r["date"], r["brand"], r["sales_amt"], r["margin"],
                                                      source="api", upload_batch_id=batch_id)
                        total += len(records)
                    except Exception:
                        pass

            # Hourly sales
            for sn in sheets:
                if "hourly" in sn.lower() and "sales" in sn.lower():
                    try:
                        result = parse_hourly_sales_file(tmp_path, sheet_name=sn)
                        records = result.get("records", [])

                        with _get_db() as conn:
                            ftype = "hourly_reference" if is_large else "hourly_sales"
                            log_upload(conn, file.filename, ftype, None,
                                       len(records), len(records), 0, None, None, None)

                            if is_large:
                                conn.execute("BEGIN")
                                for r in records:
                                    conn.execute("""
                                        INSERT OR REPLACE INTO hourly_sales
                                        (sales_site_name, sector_id, date, hour, brand, sales_amt, trans_cnt, source)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, 'api')
                                    """, (r["sales_site_name"], r.get("sector_id"), r["date"],
                                          r["hour"], r["brand"], r["sales_amt"], r["trans_cnt"]))
                                conn.execute("COMMIT")
                            else:
                                for r in records:
                                    upsert_hourly_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                                       r["date"], r["hour"], r["brand"],
                                                       r["sales_amt"], r["trans_cnt"], source="api")
                        total += len(records)
                    except Exception:
                        pass

            # ─── STORE MASTER sheets (auto-detect from combo file) ───
            storemaster_sheets = []
            for sn in sheets:
                sn_lower = sn.lower().strip()
                if sn_lower in ("store master", "store_master", "storemaster"):
                    storemaster_sheets.append(sn)
                elif "all" in sn_lower and "sites" in sn_lower:
                    storemaster_sheets.append(sn)
                elif sn_lower == "cp" and any("cmhl" in s.lower() or "sites" in s.lower() for s in sheets):
                    # CP sheet in a combo file (not a standalone blackout file)
                    storemaster_sheets.append(sn)

            if storemaster_sheets:
                try:
                    from parsers.storemaster_parser import parse_storemaster_file
                    sm_total = 0
                    for sn in storemaster_sheets:
                        try:
                            parsed = parse_storemaster_file(tmp_path)
                            # Re-parse with specific sheet
                            import pandas as _pd_sm
                            _df_sm = _pd_sm.read_excel(tmp_path, sheet_name=sn)

                            with _get_db() as conn:
                                # Detect columns
                                cols_lower = {str(c).lower().replace(" ", "").replace("_", ""): c for c in _df_sm.columns}
                                gold_col = next((cols_lower[k] for k in cols_lower if "goldcode" in k), None)
                                pos_col = next((cols_lower[k] for k in cols_lower if "poscode" in k), None)
                                cc_col = next((cols_lower[k] for k in cols_lower if "costcenter" == k.replace("name","").replace("description","").strip()), None)
                                ccn_col = next((cols_lower[k] for k in cols_lower if "costcentername" in k), None)
                                seg_col = next((cols_lower[k] for k in cols_lower if "segmentname" in k), None)
                                sector_col = next((cols_lower[k] for k in cols_lower if k == "sector"), None)
                                channel_col = next((cols_lower[k] for k in cols_lower if "channel" in k), None)
                                state_col = next((cols_lower[k] for k in cols_lower if "addressstate" in k or "address_state" in k), None)
                                township_col = next((cols_lower[k] for k in cols_lower if "addresstownship" in k or "address_township" in k), None)
                                lat_col = next((cols_lower[k] for k in cols_lower if "latitude" in k), None)
                                lon_col = next((cols_lower[k] for k in cols_lower if "longitude" in k), None)
                                size_col = next((cols_lower[k] for k in cols_lower if "storesize" in k), None)
                                open_col = next((cols_lower[k] for k in cols_lower if "opendate" in k), None)
                                close_col = next((cols_lower[k] for k in cols_lower if "closeddate" in k or "closedate" in k), None)

                                if gold_col:
                                    for _, row in _df_sm.iterrows():
                                        gc = row.get(gold_col)
                                        if pd.isna(gc):
                                            continue
                                        gc_str = str(int(gc)) if isinstance(gc, float) else str(gc).strip()
                                        cc = row.get(cc_col) if cc_col else None
                                        cc_str = str(int(cc)) if isinstance(cc, float) and cc and not pd.isna(cc) else (str(cc).strip() if cc and not pd.isna(cc) else None)
                                        ccn = str(row.get(ccn_col, '')).strip() if ccn_col and not pd.isna(row.get(ccn_col)) else ''
                                        seg = str(row.get(seg_col, '')).strip() if seg_col and not pd.isna(row.get(seg_col)) else ''
                                        sec = str(row.get(sector_col, '')).strip() if sector_col and not pd.isna(row.get(sector_col)) else ''
                                        # Map sector name to sector_id
                                        sec_map = {"retail": "CMHL", "property": "CP", "f&b": "CFC", "distribution": "PG"}
                                        sector_id = sec_map.get(sec.lower(), '') if sec else ''
                                        if not sector_id:
                                            from config.settings import SEGMENT_SECTOR_MAP
                                            sector_id = SEGMENT_SECTOR_MAP.get(seg, '')

                                        lat = row.get(lat_col) if lat_col and not pd.isna(row.get(lat_col, None)) else None
                                        lon = row.get(lon_col) if lon_col and not pd.isna(row.get(lon_col, None)) else None

                                        conn.execute("""
                                            INSERT OR REPLACE INTO store_master
                                            (gold_code, pos_code, store_name, cost_center_code, cost_center_name,
                                             segment_name, channel, address_state, address_township,
                                             latitude, longitude, store_size, open_date, closed_date, sector_id)
                                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                        """, (
                                            gc_str,
                                            str(row.get(pos_col, '')).strip() if pos_col and not pd.isna(row.get(pos_col)) else '',
                                            ccn or gc_str,
                                            cc_str,
                                            ccn,
                                            seg,
                                            str(row.get(channel_col, '')).strip() if channel_col and not pd.isna(row.get(channel_col)) else None,
                                            str(row.get(state_col, '')).strip() if state_col and not pd.isna(row.get(state_col)) else None,
                                            str(row.get(township_col, '')).strip() if township_col and not pd.isna(row.get(township_col)) else None,
                                            float(lat) if lat else None,
                                            float(lon) if lon else None,
                                            str(row.get(size_col, '')).strip() if size_col and not pd.isna(row.get(size_col)) else None,
                                            str(row.get(open_col, '')) if open_col and not pd.isna(row.get(open_col)) else None,
                                            str(row.get(close_col, '')) if close_col and not pd.isna(row.get(close_col)) else None,
                                            sector_id or None,
                                        ))
                                        sm_total += 1

                                log_upload(conn, file.filename, f"store_master_{sn}", None,
                                           sm_total, sm_total, 0, None, None, None)
                        except Exception as e:
                            import traceback; traceback.print_exc()
                    total += sm_total
                    # Enrich sites with store master data
                    try:
                        with _get_db() as conn:
                            enrich_sites_from_store_master(conn)
                    except Exception:
                        pass
                except Exception:
                    pass

            # ─── MAPPING sheet ─────────────────────────────────────
            for sn in sheets:
                if sn.lower().strip() == "mapping":
                    try:
                        import pandas as _pd_map
                        _df_map = _pd_map.read_excel(tmp_path, sheet_name=sn)
                        with _get_db() as conn:
                            for _, row in _df_map.iterrows():
                                manual = row.get(_df_map.columns[0])
                                sap_cc = row.get(_df_map.columns[1]) if len(_df_map.columns) > 1 else None
                                site_code = row.get(_df_map.columns[2]) if len(_df_map.columns) > 2 else None
                                if pd.isna(manual) or pd.isna(site_code):
                                    continue
                                # Insert into site_sales_map
                                cc_str = str(int(sap_cc)) if isinstance(sap_cc, float) and not pd.isna(sap_cc) else (str(sap_cc).strip() if sap_cc and not pd.isna(sap_cc) else '')
                                conn.execute("""
                                    INSERT OR REPLACE INTO site_sales_map
                                    (sales_site_name, site_id, sector_id, match_method)
                                    VALUES (?, ?, (SELECT sector_id FROM sites WHERE site_id = ?), 'manual')
                                """, (str(manual).strip(), str(site_code).strip(), str(site_code).strip()))
                            log_upload(conn, file.filename, "mapping", None,
                                       len(_df_map), len(_df_map), 0, None, None, None)
                    except Exception:
                        pass

            detected_type = "sales_reference" if is_large else "combo_sales"
            record_count = total

        # ─── DIESEL EXPENSE LY ─────────────────────────────────
        else:
            try:
                import pandas as _pd
                _df = _pd.read_excel(tmp_path, sheet_name=0, nrows=5)
                all_text = ' '.join(str(c).lower() for c in _df.columns)

                if any(kw in all_text for kw in ["daily average diesel", "daily normal diesel", "diesel expense", "store contribution", "store exp"]):
                    from parsers.diesel_expense_parser import parse_diesel_expense_file
                    parsed = parse_diesel_expense_file(tmp_path)
                    records = parsed.get("records", []) if isinstance(parsed, dict) else []
                    detected_type = "diesel_expense_ly"

                    with _get_db() as conn:
                        log_upload(conn, file.filename, "diesel_expense_ly", None,
                                   len(records), len(records), 0, None, None, None)
                        for r in records:
                            upsert_diesel_expense_ly(conn, r["cost_center_code"], r["sector_id"],
                                                      r["cost_center_name"],
                                                      r["yearly_expense_mmk_mil"],
                                                      r["daily_avg_expense_mmk"],
                                                      r["pct_on_sales"])
                    record_count = len(records)

                elif any(kw in all_text for kw in ["sales_amt", "sales_date", "gold_code"]):
                    from parsers.sales_parser import parse_daily_sales_file
                    parsed = parse_daily_sales_file(tmp_path)
                    records = parsed.get("records", [])
                    detected_type = "daily_sales"

                    with _get_db() as conn:
                        log_upload(conn, file.filename, "daily_sales", None,
                                   len(records), len(records), 0, None, None, None)
                        for r in records:
                            upsert_daily_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                              r["date"], r["brand"], r["sales_amt"], r["margin"],
                                              source="api")
                    record_count = len(records)
            except Exception:
                pass

        # Map sales site_id: since site_id = cost_center_code, direct match
        if detected_type in ("combo_sales", "sales_reference", "daily_sales", "hourly_sales", "hourly_reference"):
            try:
                with _get_db() as conn:
                    conn.execute("""
                        UPDATE daily_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = daily_sales.sales_site_name
                        )
                    """)
                    conn.execute("""
                        UPDATE hourly_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = hourly_sales.sales_site_name
                        )
                    """)
            except Exception:
                pass

        # Run alert checks after blackout/fuel uploads
        if detected_type and detected_type.startswith("blackout") or detected_type == "fuel_price":
            try:
                from alerts.alert_engine import run_all_checks
                run_all_checks()
            except Exception:
                pass

        return {
            "filename": file.filename,
            "sheets": sheets,
            "records": record_count,
            "type": detected_type or "unknown",
            "validation": validation if 'validation' in dir() else [],
        }
    except Exception as e:
        raise HTTPException(400, f"Parse error: {e}")
    finally:
        os.unlink(tmp_path)


# ─── 1.7 Economics ───────────────────────────────────────
@router.get("/economics")
def economics(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None, site_type: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    df = get_store_economics(date_from=date_from, date_to=date_to)
    if df.empty:
        return []
    if sector:
        df = df[df["sector_id"] == sector]
    if site_type and site_type != "All":
        df = df[df["site_type"] == site_type]
    return df.fillna("").to_dict(orient="records")


# ─── 1.8 Daily Summary ──────────────────────────────────
@router.get("/daily-summary")
def daily_summary(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None, site_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("dss.date", date_from, date_to)
        q = f"""SELECT dss.*, s.sector_id, s.site_type,
                       s.cost_center_code, s.region as site_code, s.segment_name,
                       s.cost_center_description, s.address_state, s.store_size,
                       s.latitude, s.longitude
                FROM daily_site_summary dss
                JOIN sites s ON dss.site_id = s.site_id WHERE 1=1{dsql}"""
        if sector:
            q += " AND s.sector_id = ?"; dp.append(sector)
        if site_id:
            q += " AND dss.site_id = ?"; dp.append(site_id)
        q += " ORDER BY dss.date DESC LIMIT 5000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.9 Fuel Prices ────────────────────────────────────
@router.get("/fuel-prices")
def fuel_prices(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("date", date_from, date_to)
        q = f"SELECT * FROM fuel_purchases WHERE 1=1{dsql}"
        if sector:
            q += " AND sector_id = ?"; dp.append(sector)
        q += " ORDER BY date DESC LIMIT 2000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.10 Sales ──────────────────────────────────────────
@router.get("/sales")
def sales(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    site_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("date", date_from, date_to)
        q = f"SELECT * FROM daily_sales WHERE 1=1{dsql}"
        if site_id:
            q += " AND site_id = ?"; dp.append(site_id)
        q += " ORDER BY date DESC LIMIT 5000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.11 Blackout ───────────────────────────────────────
@router.get("/blackout")
def blackout(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("do.date", date_from, date_to)
        q = f"""SELECT do.date, do.site_id, do.blackout_hr, do.gen_run_hr,
                       do.daily_used_liters, s.sector_id
                FROM daily_operations do JOIN sites s ON do.site_id = s.site_id
                WHERE do.blackout_hr IS NOT NULL{dsql}"""
        if sector:
            q += " AND s.sector_id = ?"; dp.append(sector)
        q += " ORDER BY do.date DESC LIMIT 5000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.12 Generators ────────────────────────────────────
@router.get("/generators")
def generators(
    sector: Optional[str] = None, site_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        q = """SELECT g.*, s.sector_id, s.site_type FROM generators g
               JOIN sites s ON g.site_id = s.site_id WHERE g.is_active = 1"""
        dp = []
        if sector:
            q += " AND s.sector_id = ?"; dp.append(sector)
        if site_id:
            q += " AND g.site_id = ?"; dp.append(site_id)
        q += " ORDER BY g.site_id"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.13 Site Detail ────────────────────────────────────
@router.get("/site/{site_id}")
def site_detail(
    site_id: str,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    econ = get_store_economics(date_from=date_from, date_to=date_to)
    site_row = econ[econ["site_id"] == site_id]
    if site_row.empty:
        raise HTTPException(404, f"Site {site_id} not found")

    gen_detail = get_generator_detail(site_id, date_from=date_from, date_to=date_to)
    trends = get_trends(site_id, date_from=date_from, date_to=date_to)

    return {
        "economics": site_row.fillna("").to_dict(orient="records")[0],
        "generators": gen_detail.fillna("").to_dict(orient="records") if not gen_detail.empty else [],
        "trends": {k: v.fillna("").to_dict(orient="records") if isinstance(v, pd.DataFrame) and not v.empty else [] for k, v in trends.items()} if trends else {},
    }
