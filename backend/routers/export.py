"""Export router — formatted Excel downloads matching analog terminal theme."""
import io
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.routers.auth import get_current_user

router = APIRouter()

# ─── Analog Terminal Theme ─────────────────────────────
DARK_BG = PatternFill(start_color='383832', end_color='383832', fill_type='solid')
HEADER_BG = PatternFill(start_color='EBE8DD', end_color='EBE8DD', fill_type='solid')
ROW_ALT = PatternFill(start_color='F6F4E9', end_color='F6F4E9', fill_type='solid')
ROW_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
GREEN_BG = PatternFill(start_color='007518', end_color='007518', fill_type='solid')
AMBER_BG = PatternFill(start_color='FF9D00', end_color='FF9D00', fill_type='solid')
RED_BG = PatternFill(start_color='BE2D06', end_color='BE2D06', fill_type='solid')
CYAN_BG = PatternFill(start_color='006F7C', end_color='006F7C', fill_type='solid')

TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='FEFFD6')
SUBTITLE_FONT = Font(name='Calibri', size=10, color='FEFFD6')
HEADER_FONT = Font(name='Calibri', bold=True, size=10, color='383832')
DATA_FONT = Font(name='Calibri', size=10, color='383832')
WHITE_FONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
GREEN_FONT = Font(name='Calibri', bold=True, size=10, color='007518')
RED_FONT = Font(name='Calibri', bold=True, size=10, color='BE2D06')
AMBER_FONT = Font(name='Calibri', bold=True, size=10, color='FF9D00')

BORDER = Border(
    left=Side(style='thin', color='383832'),
    right=Side(style='thin', color='383832'),
    top=Side(style='thin', color='383832'),
    bottom=Side(style='thin', color='383832'),
)
BORDER_LIGHT = Border(
    bottom=Side(style='thin', color='EBE8DD'),
)

# Status badge colors
STATUS_COLORS = {
    'OPEN': GREEN_BG, 'SAFE': GREEN_BG, 'A': GREEN_BG, 'LOW': GREEN_BG, 'STABLE': GREEN_BG,
    'MONITOR': AMBER_BG, 'WARNING': AMBER_BG, 'B': CYAN_BG, 'MEDIUM': AMBER_BG, 'WATCH': AMBER_BG,
    'REDUCE': PatternFill(start_color='F95630', end_color='F95630', fill_type='solid'),
    'CLOSE': RED_BG, 'CRITICAL': RED_BG, 'HIGH': RED_BG, 'F': RED_BG, 'DANGER': RED_BG,
    'C': AMBER_BG, 'D': PatternFill(start_color='F95630', end_color='F95630', fill_type='solid'),
}


class ExportRequest(BaseModel):
    table_name: str
    data: list[dict]
    columns: Optional[list[str]] = None
    filters: Optional[str] = None
    status_columns: Optional[list[str]] = None  # columns to apply status badge coloring


@router.post("/export/excel")
def export_excel(req: ExportRequest, user: dict = Depends(get_current_user)):
    if not req.data:
        return {"error": "No data to export"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = req.table_name[:31]  # Excel sheet name limit

    # Determine columns
    columns = req.columns or list(req.data[0].keys())
    status_cols = set(req.status_columns or [])

    # ─── Row 1: Title bar (dark background) ─────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=req.table_name.upper())
    title_cell.font = TITLE_FONT
    title_cell.fill = DARK_BG
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, len(columns) + 1):
        ws.cell(row=1, column=c).fill = DARK_BG
    ws.row_dimensions[1].height = 36

    # ─── Row 2: Subtitle (date + filters) ───────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sub = f"BCP COMMAND CENTER  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {len(req.data)} rows"
    if req.filters:
        sub += f"  |  {req.filters}"
    sub_cell = ws.cell(row=2, column=1, value=sub)
    sub_cell.font = SUBTITLE_FONT
    sub_cell.fill = DARK_BG
    sub_cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, len(columns) + 1):
        ws.cell(row=2, column=c).fill = DARK_BG
    ws.row_dimensions[2].height = 22

    # ─── Row 3: Column headers ──────────────────────────────
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name.upper().replace('_', ' '))
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[3].height = 28

    # ─── Data rows ──────────────────────────────────────────
    for row_idx, record in enumerate(req.data):
        r = row_idx + 4  # start at row 4
        is_alt = row_idx % 2 == 1
        fill = ROW_ALT if is_alt else ROW_WHITE

        for col_idx, col_name in enumerate(columns, 1):
            val = record.get(col_name, '')
            cell = ws.cell(row=r, column=col_idx, value=val if val != '' else None)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = BORDER_LIGHT
            cell.alignment = Alignment(vertical='center')

            # Number formatting
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(val, float):
                    cell.number_format = '#,##0.0' if abs(val) < 100 else '#,##0'
                else:
                    cell.number_format = '#,##0'

            # Status badge coloring
            if col_name in status_cols and isinstance(val, str):
                upper_val = val.strip().upper()
                badge_fill = STATUS_COLORS.get(upper_val)
                if badge_fill:
                    cell.fill = badge_fill
                    cell.font = WHITE_FONT
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    # ─── Auto-fit column widths ─────────────────────────────
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name) + 2
        for row_idx in range(len(req.data)):
            val = req.data[row_idx].get(col_name, '')
            if val is not None:
                max_len = max(max_len, len(str(val)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 35)

    # ─── Freeze panes (freeze header rows) ──────────────────
    ws.freeze_panes = 'A4'

    # ─── Write to buffer ────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{req.table_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
