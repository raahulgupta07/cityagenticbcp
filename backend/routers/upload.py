"""Upload router — history, store summary, raw data browser, clear actions."""
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
import pandas as pd
from utils.database import get_db
from backend.routers.auth import get_current_user, require_admin

router = APIRouter()


def _df(df):
    if df is None or (isinstance(df, pd.DataFrame) and df.empty):
        return []
    return df.fillna("").to_dict(orient="records")


# ═══════════════════════════════════════════════════════════════════════════
# Last Sync per file type
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/upload/last-sync")
def last_sync(user: dict = Depends(require_admin)):
    """Check sync status using actual DB row counts + upload history timestamps."""
    # Map card keys to DB queries that check if data exists
    checks = {
        "blackout_cmhl": ("SELECT COUNT(*) FROM daily_site_summary dss JOIN sites s ON dss.site_id = s.site_id WHERE s.sector_id = 'CMHL'", "CMHL"),
        "blackout_cp": ("SELECT COUNT(*) FROM daily_site_summary dss JOIN sites s ON dss.site_id = s.site_id WHERE s.sector_id = 'CP'", "CP"),
        "blackout_cfc": ("SELECT COUNT(*) FROM daily_site_summary dss JOIN sites s ON dss.site_id = s.site_id WHERE s.sector_id = 'CFC'", "CFC"),
        "blackout_pg": ("SELECT COUNT(*) FROM daily_site_summary dss JOIN sites s ON dss.site_id = s.site_id WHERE s.sector_id = 'PG'", "PG"),
        "fuel_price": ("SELECT COUNT(*) FROM fuel_purchases", None),
        "combo_sales": ("SELECT COUNT(*) FROM daily_sales", None),
        "diesel_expense_ly": ("SELECT COUNT(*) FROM diesel_expense_ly", None),
    }

    result = {}
    with get_db() as conn:
        for key, (query, sector) in checks.items():
            try:
                count = conn.execute(query).fetchone()[0]
            except Exception:
                count = 0

            # Get latest timestamp from upload_history or sites updated_at
            ts = None
            if key.startswith("blackout_") and sector:
                ts_row = conn.execute(
                    "SELECT MAX(updated_at) FROM sites WHERE sector_id = ?", (sector,)
                ).fetchone()
                ts = ts_row[0][:16] if ts_row and ts_row[0] else None
            if not ts:
                # Fallback: check upload_history
                ft_variants = {
                    "combo_sales": "('daily_sales','hourly_sales','combo_sales','sales','sales_reference','hourly_reference')",
                    "diesel_expense_ly": "('diesel_expense_ly','diesel_expense')",
                }
                if key in ft_variants:
                    hist = conn.execute(
                        f"SELECT uploaded_at FROM upload_history WHERE file_type IN {ft_variants[key]} ORDER BY uploaded_at DESC LIMIT 1"
                    ).fetchone()
                else:
                    hist = conn.execute(
                        "SELECT uploaded_at FROM upload_history WHERE LOWER(file_type) = LOWER(?) ORDER BY uploaded_at DESC LIMIT 1", (key,)
                    ).fetchone()
                ts = hist[0][:16] if hist and hist[0] else None

            if count > 0:
                result[key] = {"rows": count, "synced_at": ts or "seeded"}
            else:
                result[key] = None

    return result


# ═══════════════════════════════════════════════════════════════════════════
# Upload history + DB totals
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/upload/history")
def upload_history(user: dict = Depends(require_admin)):
    with get_db() as conn:
        uploads = pd.read_sql_query("""
            SELECT filename, file_type, sector_id, rows_parsed, rows_accepted,
                   rows_rejected, date_range_start, date_range_end, uploaded_at
            FROM upload_history ORDER BY uploaded_at DESC
        """, conn)
        totals = {}
        for t in ["sites", "generators", "daily_operations", "fuel_purchases",
                   "daily_sales", "hourly_sales", "store_master", "diesel_expense_ly"]:
            try:
                totals[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            except Exception:
                totals[t] = 0
    return {"uploads": _df(uploads), "totals": totals}


# ═══════════════════════════════════════════════════════════════════════════
# Store Summary (pivot: site × date × metric)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/upload/store-summary")
def store_summary(
    sector: Optional[str] = None,
    metric: str = "total_daily_used",
    user: dict = Depends(require_admin),
):
    allowed_metrics = {
        "total_gen_run_hr": "Gen Run Hours",
        "total_daily_used": "Daily Diesel Used (L)",
        "spare_tank_balance": "Tank Balance (L)",
        "blackout_hr": "Blackout Hours",
        "days_of_buffer": "Buffer Days",
        "num_generators_active": "Active Generators",
    }
    if metric not in allowed_metrics:
        raise HTTPException(400, f"Metric must be one of: {', '.join(allowed_metrics.keys())}")

    with get_db() as conn:
        q = """
            SELECT dss.date, dss.site_id, s.sector_id, dss.""" + metric + """ as value
            FROM daily_site_summary dss
            JOIN sites s ON dss.site_id = s.site_id
            WHERE 1=1
        """
        params = []
        if sector:
            q += " AND s.sector_id = ?"; params.append(sector)
        q += " ORDER BY s.sector_id, dss.site_id, dss.date"
        df = pd.read_sql_query(q, conn, params=params)

    if df.empty:
        return {"pivot": [], "dates": [], "sites": [], "metric_label": allowed_metrics[metric]}

    # Pivot: rows = site_id, columns = date
    pivot = df.pivot_table(index="site_id", columns="date", values="value", aggfunc="sum")
    dates = sorted(df["date"].unique().tolist())
    pivot = pivot.reindex(columns=dates)

    # Convert to list of dicts for frontend
    rows = []
    for site_id in pivot.index:
        row = {"site_id": site_id}
        sector_id = df[df["site_id"] == site_id]["sector_id"].iloc[0]
        row["sector_id"] = sector_id
        for d in dates:
            val = pivot.loc[site_id, d]
            row[d] = round(float(val), 1) if pd.notna(val) else None
        rows.append(row)

    return {
        "pivot": rows,
        "dates": dates,
        "sites": list(pivot.index),
        "metric": metric,
        "metric_label": allowed_metrics[metric],
    }


# ═══════════════════════════════════════════════════════════════════════════
# Raw Data Browser
# ═══════════════════════════════════════════════════════════════════════════

TABLE_QUERIES = {
    "sites": "SELECT s.site_id, s.sector_id, s.site_type, s.site_name, s.cost_center_code FROM sites s ORDER BY s.sector_id, s.site_id",
    "generators": "SELECT g.site_id, s.sector_id, g.model_name, g.power_kva, g.consumption_per_hour, g.fuel_type, g.supplier FROM generators g JOIN sites s ON g.site_id = s.site_id ORDER BY s.sector_id, g.site_id",
    "daily_ops": "SELECT do.date, do.site_id, g.model_name, do.gen_run_hr, do.daily_used_liters, do.spare_tank_balance, do.blackout_hr FROM daily_operations do JOIN generators g ON do.generator_id = g.generator_id ORDER BY do.date DESC, do.site_id LIMIT 500",
    "site_summary": "SELECT dss.date, dss.site_id, s.sector_id, dss.total_gen_run_hr, dss.total_daily_used, dss.spare_tank_balance, dss.days_of_buffer FROM daily_site_summary dss JOIN sites s ON dss.site_id = s.site_id ORDER BY dss.date DESC LIMIT 500",
    "fuel_prices": "SELECT date, sector_id, region, supplier, fuel_type, quantity_liters, price_per_liter FROM fuel_purchases ORDER BY date DESC LIMIT 200",
    "daily_sales": "SELECT date, sales_site_name, sector_id, brand, sales_amt, margin FROM daily_sales ORDER BY date DESC LIMIT 500",
    "hourly_sales": "SELECT date, hour, sales_site_name, sector_id, brand, sales_amt, trans_cnt FROM hourly_sales ORDER BY date DESC, hour LIMIT 500",
    "store_master": "SELECT gold_code, store_name, cost_center_code, segment_name, sector_id, channel, address_state FROM store_master ORDER BY sector_id LIMIT 500",
    "alerts": "SELECT alert_type, severity, site_id, sector_id, message, is_acknowledged, created_at FROM alerts ORDER BY created_at DESC LIMIT 200",
    "uploads": "SELECT filename, file_type, sector_id, rows_parsed, rows_accepted, rows_rejected, uploaded_at FROM upload_history ORDER BY uploaded_at DESC",
}


@router.get("/upload/raw/{table}")
def raw_data(
    table: str,
    search: Optional[str] = None,
    user: dict = Depends(require_admin),
):
    if table not in TABLE_QUERIES:
        raise HTTPException(400, f"Table must be one of: {', '.join(TABLE_QUERIES.keys())}")

    with get_db() as conn:
        df = pd.read_sql_query(TABLE_QUERIES[table], conn)

    if search and not df.empty:
        mask = df.astype(str).apply(lambda row: row.str.contains(search, case=False).any(), axis=1)
        df = df[mask]

    return {"rows": _df(df), "count": len(df), "table": table}


# ═══════════════════════════════════════════════════════════════════════════
# Data Quality Validation — Excel vs DB comparison per sector
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/upload/validation")
def data_validation(user: dict = Depends(get_current_user)):
    """Compare DB totals against Excel source file totals for all uploaded blackout files."""
    import glob, os
    from parsers.blackout_parser import parse_blackout_file
    from parsers.base_parser import parse_date_from_cell, clean_numeric
    import openpyxl

    sectors = []

    with get_db() as conn:
        # Get last uploaded file path per sector from upload_history
        for sector_id in ["CFC", "CMHL", "CP", "PG"]:
            # Check if sector has data
            count = conn.execute(
                "SELECT COUNT(*) FROM daily_site_summary dss JOIN sites s ON s.site_id=dss.site_id WHERE s.sector_id=?",
                (sector_id,)
            ).fetchone()[0]
            if count == 0:
                continue

            # Get DB totals per date
            db_rows = conn.execute(f"""
                SELECT dss.date,
                    ROUND(SUM(dss.total_gen_run_hr),1) as gen_hr,
                    ROUND(SUM(dss.total_daily_used),1) as fuel,
                    ROUND(SUM(dss.spare_tank_balance),0) as tank,
                    ROUND(SUM(dss.blackout_hr),1) as blackout,
                    COUNT(DISTINCT dss.site_id) as sites
                FROM daily_site_summary dss
                JOIN sites s ON s.site_id = dss.site_id
                WHERE s.sector_id = ?
                GROUP BY dss.date ORDER BY dss.date
            """, (sector_id,)).fetchall()

            # Get site/gen counts
            n_sites = conn.execute("SELECT COUNT(*) FROM sites WHERE sector_id=?", (sector_id,)).fetchone()[0]
            n_gens = conn.execute(
                "SELECT COUNT(*) FROM generators WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id=?)",
                (sector_id,)
            ).fetchone()[0]

            # Try to find the source Excel file
            last_upload = conn.execute(
                "SELECT filename FROM upload_history WHERE file_type=? ORDER BY uploaded_at DESC LIMIT 1",
                (f"blackout_{sector_id.lower()}",)
            ).fetchone()

            excel_totals = {}
            date_cols_list = []
            fpath = None
            if last_upload:
                # Search for file in common locations
                fname = last_upload[0]
                search_paths = [
                    f"/Users/rahulgupta/Desktop/CityAI-Final-Project/CityBCPAgent/Data2/{fname}",
                    f"/Users/rahulgupta/Desktop/CityAI-Final-Project/CityBCPAgent/Data/{fname}",
                ]
                # Also search by pattern
                for pattern_dir in ["/Users/rahulgupta/Desktop/CityAI-Final-Project/CityBCPAgent/Data2",
                                     "/Users/rahulgupta/Desktop/CityAI-Final-Project/CityBCPAgent/Data"]:
                    if os.path.isdir(pattern_dir):
                        for f in os.listdir(pattern_dir):
                            if sector_id.lower() in f.lower() and "blackout" in f.lower() and f.endswith(".xlsx"):
                                search_paths.append(os.path.join(pattern_dir, f))

                fpath = None
                for sp in search_paths:
                    if os.path.exists(sp):
                        fpath = sp
                        break

                if fpath:
                    try:
                        parsed = parse_blackout_file(fpath, sector_id)
                        cols_det = parsed.get("columns_detected", {})
                        dsub = cols_det.get("date_sub_offsets", {})
                        og = dsub.get("gen_run_hr", 0)
                        of_ = dsub.get("daily_used_liters", 1)
                        ot = dsub.get("spare_tank_balance", 2)
                        ob = dsub.get("blackout_hr", 3)

                        wb = openpyxl.load_workbook(fpath, data_only=True)
                        ws = wb[sector_id] if sector_id in wb.sheetnames else wb[wb.sheetnames[0]]

                        total_row = None
                        for r in range(1, 15):
                            for c in range(1, 8):
                                v = ws.cell(row=r, column=c).value
                                if v and "total" in str(v).lower():
                                    total_row = r
                                    break
                            if total_row:
                                break

                        if total_row:
                            for ci in range(1, ws.max_column + 1):
                                ds = parse_date_from_cell(ws.cell(row=2, column=ci).value)
                                if ds:
                                    date_cols_list.append((ds, ci))
                                    excel_totals[ds] = {
                                        "gen_hr": clean_numeric(ws.cell(row=total_row, column=ci + og).value) or 0,
                                        "fuel": clean_numeric(ws.cell(row=total_row, column=ci + of_).value) or 0,
                                        "tank": clean_numeric(ws.cell(row=total_row, column=ci + ot).value) or 0,
                                        "blackout": clean_numeric(ws.cell(row=total_row, column=ci + ob).value) or 0,
                                    }
                        wb.close()
                    except Exception:
                        pass

            # Detect Excel issues (time-formatted cells, missing formulas)
            issues = []
            if fpath:
                try:
                    from datetime import time as _dt_time
                    _wb2 = openpyxl.load_workbook(fpath, data_only=True)
                    _ws2 = _wb2[sector_id] if sector_id in _wb2.sheetnames else _wb2[_wb2.sheetnames[0]]

                    # Check for missing blackout formula
                    if total_row and date_cols_list:
                        _first_bo_col = date_cols_list[0][1] + ob
                        _bo_formula_val = _ws2.cell(row=total_row, column=_first_bo_col).value
                        if _bo_formula_val is None or _bo_formula_val == 0:
                            # Check if there's actual blackout data
                            _has_bo_data = False
                            for _r in range(total_row + 1, min(_ws2.max_row + 1, 80)):
                                _bv = _ws2.cell(row=_r, column=_first_bo_col).value
                                if _bv is not None and _bv != 0:
                                    _has_bo_data = True; break
                            if _has_bo_data:
                                issues.append({"type": "NO_FORMULA", "col": "Blackout", "msg": "Excel total row has no SUM formula for Blackout — total shows 0 but data exists"})

                    # Check for time-formatted blackout cells
                    _time_sites = {}
                    for _ds, _dc in date_cols_list:
                        _bo_col = _dc + ob
                        for _r in range(total_row + 1 if total_row else 6, min(_ws2.max_row + 1, 80)):
                            _raw = _ws2.cell(row=_r, column=_bo_col).value
                            if isinstance(_raw, _dt_time) and (_raw.hour > 0 or _raw.minute > 0):
                                _cc = _ws2.cell(row=_r, column=6).value
                                _name = _ws2.cell(row=_r, column=5).value or _ws2.cell(row=_r, column=4).value
                                if _cc:
                                    _cc_str = str(int(_cc)) if isinstance(_cc, float) else str(_cc)
                                    if _cc_str not in _time_sites:
                                        _time_sites[_cc_str] = {"name": str(_name), "dates": 0}
                                    _time_sites[_cc_str]["dates"] += 1

                    for _cc, _info in sorted(_time_sites.items()):
                        issues.append({"type": "TIME_FORMAT", "col": "Blackout", "site_id": _cc, "site_name": _info["name"],
                                       "msg": f"{_info['name']} ({_cc}): {_info['dates']} cells use h:mm time format — Excel SUM treats as day-fraction, not hours"})

                    _wb2.close()
                except Exception:
                    pass

            # Build comparison rows
            rows = []
            for db_row in db_rows:
                d = db_row[0]
                db = {"gen_hr": db_row[1] or 0, "fuel": db_row[2] or 0, "tank": db_row[3] or 0, "blackout": db_row[4] or 0}
                e = excel_totals.get(d, {"gen_hr": 0, "fuel": 0, "tank": 0, "blackout": 0})
                has_excel = d in excel_totals
                # For blackout: if no formula or time-format issues, don't fail
                bo_issue = any(i["col"] == "Blackout" for i in issues)
                pass_check = all(abs(e[k] - db[k]) < 2 for k in (["gen_hr", "fuel", "tank", "blackout"] if not bo_issue else ["gen_hr", "fuel", "tank"])) if has_excel else True
                rows.append({
                    "date": d,
                    "sites": db_row[5],
                    "excel": {k: round(e[k], 1) for k in ["gen_hr", "fuel", "tank", "blackout"]},
                    "db": {k: round(db[k], 1) for k in ["gen_hr", "fuel", "tank", "blackout"]},
                    "pass": pass_check,
                    "has_excel": has_excel,
                    "bo_issue": bo_issue,
                })

            sectors.append({
                "sector": sector_id,
                "sites": n_sites,
                "generators": n_gens,
                "rows": rows,
                "issues": issues,
            })

    return {"sectors": sectors}


# ═══════════════════════════════════════════════════════════════════════════
# Clear data
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/upload/clear/{target}")
def clear_data(target: str, user: dict = Depends(require_admin)):
    def _sector_clear(sec):
        """Full cleanup for a sector: operations → summary → generators → sites."""
        return [
            f"DELETE FROM daily_operations WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='{sec}')",
            f"DELETE FROM daily_site_summary WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='{sec}')",
            f"DELETE FROM generators WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='{sec}')",
            f"DELETE FROM sites WHERE sector_id='{sec}'",
        ]
    CLEAR_MAP = {
        "all": [
            "DELETE FROM daily_operations", "DELETE FROM daily_site_summary",
            "DELETE FROM generators", "DELETE FROM sites",
            "DELETE FROM fuel_purchases", "DELETE FROM daily_sales",
            "DELETE FROM hourly_sales", "DELETE FROM alerts", "DELETE FROM ai_insights_cache",
        ],
        "cmhl": _sector_clear("CMHL"),
        "cp": _sector_clear("CP"),
        "cfc": _sector_clear("CFC"),
        "pg": _sector_clear("PG"),
        "fuel": ["DELETE FROM fuel_purchases"],
        "sales": ["DELETE FROM daily_sales", "DELETE FROM hourly_sales"],
    }
    if target not in CLEAR_MAP:
        raise HTTPException(400, f"Target must be one of: {', '.join(CLEAR_MAP.keys())}")

    with get_db() as conn:
        for sql in CLEAR_MAP[target]:
            conn.execute(sql)
        conn.execute("DELETE FROM alerts")
        conn.execute("DELETE FROM ai_insights_cache")

    return {"ok": True, "message": f"Cleared {target} data"}
