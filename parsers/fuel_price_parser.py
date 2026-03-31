"""
Parser for Daily Fuel Price.xlsx (4 sheets: CMHL, CP, CFC, PG).

Each sheet layout:
  Row 1: Title
  Row 2: Region markers (YGN cols 3-8, MDY cols 9-14)
  Row 3: Headers (Sector, Date, Supplier, Qty, Price PD, Supplier, Qty, Price HSD, ...)
  Row 4+: Data
"""
import openpyxl

from parsers.base_parser import clean_numeric, parse_date_from_cell, clean_value


def parse_fuel_price_file(filepath):
    """
    Parse Daily Fuel Price.xlsx and return structured data for all sheets.

    Returns:
        {
            "purchases": [
                {
                    "sector_id": str, "date": str, "region": str,
                    "supplier": str, "fuel_type": str,
                    "quantity_liters": float, "price_per_liter": float,
                }
            ],
            "errors": [str],
            "warnings": [str],
        }
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    result = {
        "purchases": [],
        "errors": [],
        "warnings": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sector_id = sheet_name.strip().upper()
        if sector_id not in ("CP", "CMHL", "CFC", "PG"):
            result["warnings"].append(f"Unknown sheet '{sheet_name}', skipping")
            continue

        # Layout:
        # Col 1: Sector
        # Col 2: Date
        # YGN region:
        #   Col 3: Supplier (PD)
        #   Col 4: Qty (L)
        #   Col 5: Price PD
        #   Col 6: Supplier (HSD)
        #   Col 7: Qty (L)
        #   Col 8: Price HSD
        # MDY region:
        #   Col 9: Supplier (PD)
        #   Col 10: Qty (L)
        #   Col 11: Price PD
        #   Col 12: Supplier (HSD)
        #   Col 13: Qty (L)
        #   Col 14: Price HSD

        # Find data start row (first row where col 2 has a date)
        data_start = None
        for r in range(3, min(10, ws.max_row + 1)):
            date_val = parse_date_from_cell(ws.cell(row=r, column=2).value)
            if date_val:
                data_start = r
                break

        if data_start is None:
            # Try row 4 as fallback
            data_start = 4

        for row_idx in range(data_start, ws.max_row + 1):
            date_str = parse_date_from_cell(ws.cell(row=row_idx, column=2).value)
            if not date_str:
                continue

            # YGN PD (cols 3, 4, 5)
            _add_purchase(result, sector_id, date_str, "YGN", "PD",
                          ws.cell(row=row_idx, column=3).value,
                          ws.cell(row=row_idx, column=4).value,
                          ws.cell(row=row_idx, column=5).value)

            # YGN HSD (cols 6, 7, 8)
            _add_purchase(result, sector_id, date_str, "YGN", "HSD",
                          ws.cell(row=row_idx, column=6).value,
                          ws.cell(row=row_idx, column=7).value,
                          ws.cell(row=row_idx, column=8).value)

            # MDY PD (cols 9, 10, 11)
            _add_purchase(result, sector_id, date_str, "MDY", "PD",
                          ws.cell(row=row_idx, column=9).value,
                          ws.cell(row=row_idx, column=10).value,
                          ws.cell(row=row_idx, column=11).value)

            # MDY HSD (cols 12, 13, 14)
            _add_purchase(result, sector_id, date_str, "MDY", "HSD",
                          ws.cell(row=row_idx, column=12).value,
                          ws.cell(row=row_idx, column=13).value,
                          ws.cell(row=row_idx, column=14).value)

    wb.close()
    return result


def _add_purchase(result, sector_id, date_str, region, fuel_type,
                  supplier_val, qty_val, price_val):
    """Add a fuel purchase record if there's meaningful data."""
    supplier = clean_value(supplier_val)
    quantity = clean_numeric(qty_val)
    price = clean_numeric(price_val)

    # Skip if no supplier AND no price (completely empty entry)
    if supplier is None and price is None:
        return

    # Clean supplier name
    if isinstance(supplier, (int, float)):
        supplier = None  # numeric in supplier column = data error
    if supplier:
        supplier = str(supplier).strip()

    result["purchases"].append({
        "sector_id": sector_id,
        "date": date_str,
        "region": region,
        "supplier": supplier,
        "fuel_type": fuel_type,
        "quantity_liters": quantity,
        "price_per_liter": price,
    })
